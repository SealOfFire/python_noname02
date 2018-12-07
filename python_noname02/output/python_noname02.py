#coding:utf-8
from LogHandle import Logger
import urllib3
#import xlrd
#from xlutils.copy import copy
import os
import sys
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

Logger.debug(u'全局变量声明')
Logger.debug(u'声明http')
http = urllib3.PoolManager()

#r = http.request('GET', 'https://www.baidu.com')
#Logger.info("http status code:[%s]" % r.status) # 200
def getHttpStatusCode(url):
	'''
	获取http状态吗
	'''
	Logger.debug(u'---- getHttpStatusCode ---- BEGIN ----')
	Logger.debug(u'---- params1:url:%s ----' % url)

	# url为空时直接返回
	Logger.debug(u'---- return ----')
	if(url == None or url == ''): return ""
	#
	r = http.request('GET', url)
	Logger.info("%s : %s" % (url, r.status))

	Logger.debug(u'---- return1:%s ----' % r.status)
	Logger.debug(u'---- getHttpStatusCode ---- END ----')
	return r.status


def getExcelFilesPath():
	'''
	获取当前文件夹下的所有excel文件
	'''
	Logger.debug(u'---- getExcelFilesPath ---- BEGIN ----')
	
	Logger.debug(u'获取当前执行的文件夹路径')
	homePath = os.path.dirname(os.path.realpath(__file__))
	Logger.info(u'当前路径:"%s"' % homePath)
	
	Logger.debug(u'遍历路径下的文件')
	list = os.listdir(homePath)
	excelFiles = []
	for i in range(0,len(list)):
		filePath = os.path.join(homePath,list[i])
		if (os.path.isfile(filePath) and os.path.splitext(filePath)[1] in ['.xls', '.xlsx']):
			Logger.info(u'excel文件:"%s"' % filePath)
			excelFiles.append(filePath)

	Logger.debug(u'---- return 总共%s个excel文件 ----' % len(excelFiles))
	Logger.debug(u'---- getExcelFilesPath ---- END ----')
	return excelFiles


def readExcel(path):
	'''
	读取excel
	'''
	Logger.debug(u'---- readExcel ---- BEGIN ----')
	Logger.debug(u'---- params1:path:%s ----' % path)
	
	Logger.info(u'打开excel:"%s"' % path)
	workbook = load_workbook(filename = path)
	Logger.debug(u'查看所有sheet页')
	for sheetName in workbook.sheetnames:
		Logger.info(u'开始处理sheet:%s' % sheetName)
		#sheet = workbook.sheet_by_name(sheetName)
		#sheetnew = workbooknew.get_sheet(0)
		sheet = workbook[sheetName]
		for row in sheet.iter_rows(min_row=2):
			Logger.debug(u'获取url')
			url1 = row[6].value
			url2 = row[7].value
			url3 = row[8].value
			Logger.debug(u'获取url访问结果')
			status1 = ''
			status2 = ''
			status3 = ''
			today1 = ''
			today2 = ''
			today3 = ''
			if(url1 != None and url1 != ''):
				today1 = datetime.datetime.today()
				status1 = getHttpStatusCode(url1)

			if(url2 != None and url2 != ''):
				status2 = getHttpStatusCode(url2)
				today2 = datetime.datetime.today()
			
			if(url3 != None and url3 != ''):
				status3 = getHttpStatusCode(url3)
				today3 = datetime.datetime.today()

			Logger.debug(u'访问结果写回到excel')
			# 结果写回excel
			index = 11;
			row[12].value = today1
			row[13].value = status1
			row[14].value = today2
			row[15].value = status2
			row[16].value = today3
			row[17].value = status3

		Logger.debug(u'保存excel的修改')
		workbook.save(path)
	Logger.debug(u'---- readExcel ---- END ----')
	pass


Logger.info(u'------------ 程序开始 ------------')
if __name__ == "__main__":
	try:
		excelFiles = getExcelFilesPath()
		for file in excelFiles:
			Logger.info(u'处理:"%s"' % file)
			readExcel(file)
		
		Logger.info(u'------------ 程序完成 正常 ------------')
	except Exception as e:
		Logger.error(e)
		Logger.info(u'------------ 程序完成 异常------------')
