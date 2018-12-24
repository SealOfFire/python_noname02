#coding:utf-8
from LogHandle import Logger
import urllib3
import certifi
#import xlrd
#from xlutils.copy import copy
import os
import sys
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

Logger.debug(u'全局变量声明')
Logger.debug(u'声明http')
http = urllib3.PoolManager(cert_reqs='CERT_REQUIRED', ca_certs=certifi.where())
# 结果开始写入的索引号
resultIndex = 11
# 开始读取的行
startReadRowIndex = 2
# 读取url的位置索引
urlIndex = [6,7,8]


#r = http.request('GET', 'https://www.baidu.com')
#Logger.info("http status code:[%s]" % r.status) # 200
def getHttpStatusCode(url):
	'''
	获取http状态吗
	'''
	Logger.debug(u'---- getHttpStatusCode ---- BEGIN ----')
	Logger.debug(u'---- params1:url:%s ----' % url)

	# url为空时直接返回
	if(url == None or url == ''): 
		Logger.debug(u'---- return ----')
		return ""
	#
	try:
		r = http.request('GET', url)
		Logger.info("%s : %s" % (url, r.status))
	except urllib3.exceptions.MaxRetryError as e:
		Logger.debug(u'---- return ----')
		Logger.debug(e)
		return '链接无效'
	except:
		raise
	Logger.debug(u'---- return1:%s ----' % r.status)
	Logger.debug(u'---- getHttpStatusCode ---- END ----')
	return r.status


def getHttpStatusCode2(url, result):
	'''
	获取http状态吗
	'''
	Logger.debug(u'---- getHttpStatusCode2 ---- BEGIN ----')
	Logger.debug(u'---- params1:url:%s ----' % url)
	status = ''
	try:
		r = http.urlopen(method='GET', url=url, timeout=10, redirect=False)
		status = r.status
		result.append([url,status])
		Logger.info("%s : %s" % (url, status))
		host = urllib3.get_host(url)
		if(status in [301,302]):
			redirect = r.get_redirect_location()
			if(urllib3.get_host(redirect)[1] == None):
				redirect = host[0] + '://' + host[1] + '/' + redirect
			Logger.debug(u'重定向url:%s' % redirect)
			return getHttpStatusCode2(redirect,result)
	except urllib3.exceptions.MaxRetryError as e:
		Logger.debug(u'---- return ----')
		result.append([url,'链接无效'])
		return '链接无效'
	except:
		raise
	else:
		return status
		Logger.debug(u'---- return1:%s ----' % status)
		Logger.debug(u'---- getHttpStatusCode2 ---- END ----')
def getExcelFilesPath():
	'''
	获取当前文件夹下的所有excel文件
	'''
	Logger.debug(u'---- getExcelFilesPath ---- BEGIN ----')
	
	Logger.debug(u'获取当前执行的文件夹路径')
	homePath = os.path.dirname(os.path.realpath(__file__))
	Logger.info(u'当前路径:"%s"' % homePath)
	
	Logger.debug(u'遍历路径下的文件')
	list = os.listdir(homePath)
	excelFiles = []
	for i in range(0,len(list)):
		filePath = os.path.join(homePath,list[i])
		if (os.path.isfile(filePath) and os.path.splitext(filePath)[1] in ['.xls', '.xlsx']):
			Logger.info(u'excel文件:"%s"' % filePath)
			excelFiles.append(filePath)

	Logger.debug(u'---- return 总共%s个excel文件 ----' % len(excelFiles))
	Logger.debug(u'---- getExcelFilesPath ---- END ----')
	return excelFiles


def getUrls(row):
	'''
	获取要访问的地址列表
	'''
	urls = []
	for i in urlIndex:
		urls.append(row[i].value)
	return urls


def writeExcel(result, row):
	'''
	结果写回到excel
	'''
	index = resultIndex #开始写入列的序号
	for r in result:
		row[index].value = r[0] # 时间
		index+=1
		row[index].value = r[1]  # 结果
		index+=1


def readExcel(path):
	'''
	读取excel
	'''
	Logger.debug(u'---- readExcel ---- BEGIN ----')
	Logger.debug(u'---- params1:path:%s ----' % path)
	
	Logger.info(u'打开excel:"%s"' % path)
	workbook = load_workbook(filename = path)
	Logger.debug(u'查看所有sheet页')
	for sheetName in workbook.sheetnames:
		Logger.info(u'开始处理sheet:%s' % sheetName)
		#sheet = workbook.sheet_by_name(sheetName)
		#sheetnew = workbooknew.get_sheet(0)
		sheet = workbook[sheetName]
		for row in sheet.iter_rows(min_row=startReadRowIndex):
			Logger.debug(u'获取url')
			urls = getUrls(row)

			Logger.debug(u'获取url访问结果')
			result = []
			for url in urls:
				if(url != None and url != ''):
					today = datetime.datetime.today()
					urlRedirectStack = []
					status = getHttpStatusCode2(url,urlRedirectStack)
					status = urlRedirectStack[0][1]
					result.append([today,status])
				else:
					result.append(['',''])

			Logger.debug(u'访问结果写回到excel')
			# 结果写回excel
			writeExcel(result, row)

			# 每行修改保存一次
			Logger.debug(u'保存excel的修改')
			workbook.save(path)
	Logger.debug(u'---- readExcel ---- END ----')


Logger.info(u'------------ 程序开始 ------------')
if __name__ == "__main__":
	try:
		excelFiles = getExcelFilesPath()
		for file in excelFiles:
			Logger.info(u'处理:"%s"' % file)
			readExcel(file)
	except Exception as e:
		Logger.error(e)
		Logger.info(u'------------ 程序完成 异常------------')
	else:
		Logger.info(u'------------ 程序完成 正常 ------------')
